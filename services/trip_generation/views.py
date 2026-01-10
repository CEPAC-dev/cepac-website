from django.http import JsonResponse
from django.shortcuts import render
import requests
import json
from openpyxl.utils import get_column_letter
import re
import zipfile
from io import BytesIO
from collections import defaultdict
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side










def index(request):
    """صفحة البداية لخدمة توليد الرحلات"""
    return render(request, 'trip_generation/index.html')

def showReport(request):
    """الواجهة الرئيسية"""
    return render(request, 'trip_generation/showReport.html')

def tripGeneration(request):
    """عرض تقرير توليد الرحلات"""
    return render(request, 'trip_generation/tripGeneration.html')



# =====================================================
# Helpers
# =====================================================

def safe_sheet_title(title):
    """Ensure the Excel sheet title is valid and <= 31 chars."""
    if not title:
        return "Sheet"
    title = re.sub(r'[\\/*?:\[\]]', '-', title)
    return title.strip()[:31] or "Sheet"


def merge_same_cells(ws, col_index, start_row=2):
    """Merge consecutive cells with the same value in a column."""
    current_value = ws.cell(row=start_row, column=col_index).value
    merge_start = start_row

    for row in range(start_row + 1, ws.max_row + 2):
        value = ws.cell(row=row, column=col_index).value if row <= ws.max_row else None

        if value != current_value:
            if merge_start < row - 1:
                ws.merge_cells(
                    start_row=merge_start,
                    start_column=col_index,
                    end_row=row - 1,
                    end_column=col_index
                )
                ws.cell(row=merge_start, column=col_index).alignment = Alignment(horizontal="center", vertical="center")

            current_value = value
            merge_start = row


# =====================================================
# Styles
# =====================================================

CENTER = Alignment(horizontal="center", vertical="center")
HEADER_FILL = PatternFill("solid", fgColor="2F80ED")
HEADER_FONT = Font(color="FFFFFF", bold=True)

thin = Side(style="thin")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

PERIOD_COLORS = {
    "AM": PatternFill("solid", fgColor="D6E4FF"),
    "MD": PatternFill("solid", fgColor="E6F4EA"),
    "PM": PatternFill("solid", fgColor="FFF4E5"),
    "PHG": PatternFill("solid", fgColor="F3E8FF"),
}


# =====================================================
# Main View
# =====================================================

@csrf_exempt
def export_trip_generation_excel(request):
    data = json.loads(request.body)
    results = data.get("results", [])

    # ---------- Group by Manual ----------
    manuals = defaultdict(list)
    for item in results:
        manual_name = item.get("manual", {}).get("name", "Manual")
        manuals[manual_name].append(item)

    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:

        # لكل Manual ملف Excel
        for manual_name, manual_items in manuals.items():

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # ---------- Group by Category (for sheet name) ----------
            categories = defaultdict(list)
            for item in manual_items:
                category = item.get("category", {}).get("name", "Category")
                categories[category].append(item)

            # =================================================
            # Create Sheets by Category
            # =================================================
            for category_name, items in categories.items():

                sheet_title = safe_sheet_title(category_name)
                ws = wb.create_sheet(title=sheet_title)

                # ---------- Header ----------
                headers = [
                    "Category",
                    "Sub Category",
                    "Sub Sub Category",
                    "Independent Variable",
                    "Day",
                    "Period",
                    "Rate",
                    "In",
                    "Out",
                ]
                ws.append(headers)

                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = CENTER
                    cell.border = BORDER

                # ---------- Data ----------
                for item in items:
                    for rate in item.get("rates", []):

                        if not isinstance(rate.get("rate"), (int, float)):
                            continue

                        ws.append([
                            item.get("category", {}).get("name", ""),
                            item.get("subCategory", {}).get("name", ""),
                            item.get("subSubCategory", {}).get("name", ""),
                            item.get("independentVariable", ""),
                            rate.get("day", ""),
                            rate.get("period", ""),
                            rate.get("rate", 0),
                            rate.get("in", 0),
                            rate.get("out", 0),
                        ])

                        row = ws.max_row
                        period = str(rate.get("period", "")).upper()
                        fill = PERIOD_COLORS.get(period)

                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.alignment = CENTER
                            cell.border = BORDER
                            if fill:
                                cell.fill = fill

                # ---------- Merge ----------
                merge_same_cells(ws, 1)
                merge_same_cells(ws, 2)
                merge_same_cells(ws, 3)
                merge_same_cells(ws, 4)

                # ---------- Freeze & Filter ----------
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = f"A1:I{ws.max_row}"

            # ---------- Save Excel ----------
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            safe_name = re.sub(r'[\\/*?:\[\]]', '-', manual_name)
            zip_file.writestr(f"{safe_name}.xlsx", excel_buffer.read())

    zip_buffer.seek(0)

    response = HttpResponse(zip_buffer, content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="Trip_Generation_Reports.zip"'
    return response
