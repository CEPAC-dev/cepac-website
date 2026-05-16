import os
import uuid
import base64
import pickle
from io import BytesIO
from django.shortcuts import render, redirect
from django.http import HttpResponse, FileResponse, Http404
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as XLImage

# Base temporary folder
TMP_DIR_BASE = "services/outlier/media"
os.makedirs(TMP_DIR_BASE, exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# Green shades per iteration (light → dark).
# Add more entries if you ever need more than 6 iterations.
# ─────────────────────────────────────────────────────────────────────────────
ITERATION_COLORS = [
    "C6EFCE",   # Iteration 1 – light green
    "70AD47",   # Iteration 2 – medium green
    "375623",   # Iteration 3 – dark green
    "A9D18E",   # Iteration 4 – soft green
    "538135",   # Iteration 5 – forest green
    "243F00",   # Iteration 6 – very dark green
]


@login_required
def outlier_detection(request):
    """Initial upload and processing route"""
    if request.method == "POST":
        if "file_session_id" in request.POST and "processing_mode" in request.POST:
            return process_with_mode(request)

        if "file" not in request.FILES:
            return render(request, "outlier/upload.html", {"error": "Please choose a file."})

        excel_file = request.FILES["file"]

        try:
            if excel_file.name.endswith(".csv"):
                df = pd.read_csv(excel_file)
            else:
                df = pd.read_excel(excel_file)
        except Exception as e:
            return render(request, "outlier/upload.html", {"error": f"Failed to read file: {e}"})

        rate_columns = [col for col in df.columns if 'rate' in col.lower()]
        if not rate_columns:
            return render(request, "outlier/upload.html", {"error": "No 'Rate' columns found."})

        session_id   = str(uuid.uuid4().hex)
        session_dir  = os.path.join(TMP_DIR_BASE, str(request.user.id), "sessions")
        os.makedirs(session_dir, exist_ok=True)
        session_file = os.path.join(session_dir, f"{session_id}.pkl")

        with open(session_file, 'wb') as f:
            pickle.dump(df, f)

        preview_data = df.head(5).to_dict('records')
        context = {
            "file_session_id": session_id,
            "columns":         list(df.columns),
            "rate_columns":    rate_columns,
            "preview_data":    preview_data,
            "total_rows":      len(df)
        }

        return render(request, "outlier/select_mode.html", context)

    return render(request, "outlier/upload.html")


def process_with_mode(request):
    """Process data based on selected mode"""
    session_id      = request.POST.get("file_session_id")
    processing_mode = request.POST.get("processing_mode")

    selected_items = []
    sel_final = request.POST.get("selected_items_final")
    if sel_final:
        selected_items = [s.strip() for s in sel_final.split(",") if s.strip()]
    else:
        selected_items = request.POST.getlist("selected_items")

    session_dir  = os.path.join(TMP_DIR_BASE, str(request.user.id), "sessions")
    session_file = os.path.join(session_dir, f"{session_id}.pkl")

    if not os.path.exists(session_file):
        return render(request, "outlier/upload.html",
                      {"error": "Session expired. Please upload again."})

    with open(session_file, 'rb') as f:
        df = pickle.load(f)

    user_tmp_dir = os.path.join(TMP_DIR_BASE, str(request.user.id), str(uuid.uuid4().hex))
    os.makedirs(user_tmp_dir, exist_ok=True)

    results = {}

    if processing_mode == "single_column":
        col = selected_items[0]
        results[col] = process_column(df, col, user_tmp_dir)

    elif processing_mode == "multiple_columns":
        for col in selected_items:
            results[col] = process_column(df, col, user_tmp_dir)

    elif processing_mode == "single_row":
        row_num = int(selected_items[0]) - 1
        results = process_rows(df, [row_num], user_tmp_dir, processing_mode)

    elif processing_mode == "multiple_rows":
        row_nums = [int(x) - 1 for x in selected_items]
        results  = process_rows(df, row_nums, user_tmp_dir, processing_mode)

    if not results:
        return render(request, "outlier/upload.html", {"error": "No data to process."})

    excel_filename = f"outlier_results_{session_id}.xlsx"
    excel_path     = os.path.join(user_tmp_dir, excel_filename)
    create_results_excel(df, results, excel_path, processing_mode, selected_items)

    return render(request, "outlier/result.html", {
        "results":         results,
        "selected_items":  selected_items,
        "excel_filename":  excel_filename,
        "tmp_folder":      user_tmp_dir,
        "processing_mode": processing_mode
    })


def process_column(df, rate_col, user_tmp_dir):
    """Process a single column for outliers — iterative removal"""
    data = pd.to_numeric(df[rate_col], errors="coerce").dropna()
    if len(data) < 3:
        return None

    iteration       = 0
    current         = data.copy()
    global_outliers = pd.Series(False, index=data.index)
    # Maps each outlier's DataFrame index → 1-based iteration number
    outlier_iteration_map: dict = {}
    summary_rows = []

    while True:
        mean_val = current.mean()
        std_val  = current.std()

        outliers     = (current < mean_val - std_val) | (current > mean_val + std_val)
        num_outliers = outliers.sum()

        summary_rows.append([
            f"Iteration {iteration + 1}",
            round(float(mean_val), 4),
            round(float(std_val),  4),
            int(num_outliers)
        ])

        if num_outliers == 0:
            break

        newly_found = current[outliers].index
        global_outliers.loc[newly_found] = True

        for idx in newly_found:
            outlier_iteration_map[idx] = iteration + 1   # 1-based

        current   = current[~outliers]
        iteration += 1

    # ── Plot ──────────────────────────────────────────────────────────────────
    plt.figure(figsize=(9, 4))

    normal_points  = data[~global_outliers]
    outlier_points = data[global_outliers]

    plt.scatter(normal_points.index, normal_points,
                color="blue", label="Data", alpha=0.7, s=60)
    for i, val in normal_points.items():
        plt.text(i, val, f"{val}", color="blue", fontsize=9, ha='center', va='bottom')

    plt.scatter(outlier_points.index, outlier_points,
                color="red", label="Outliers", s=120,
                edgecolors="black", linewidth=1.5)
    for i, val in outlier_points.items():
        plt.text(i, val, f"{val}", color="red", fontsize=9, ha='center', va='bottom')

    plt.axhline(y=current.mean(), color="green", linestyle="--",
                label="Mean (Final)", linewidth=2)
    plt.legend()
    plt.title(f"Outlier Detection - {rate_col}")
    plt.xlabel("Index")
    plt.ylabel(rate_col)

    buffer = BytesIO()
    plt.savefig(buffer, format="png", dpi=100, bbox_inches='tight')
    buffer.seek(0)
    plot_data = base64.b64encode(buffer.getvalue()).decode()
    plt.close()

    cleaned_total      = len(current)
    total              = len(data)
    total_outliers     = total - cleaned_total
    outlier_percentage = (total_outliers / total) * 100 if total > 0 else 0

    return {
        "cleaned_total":         cleaned_total,
        "total":                 total,
        "total_outliers":        total_outliers,
        "outlier_percentage":    round(outlier_percentage, 2),
        "final_mean":            round(float(current.mean()), 4) if cleaned_total > 0 else "N/A",
        "final_std":             round(float(current.std()),  4) if cleaned_total > 0 else "N/A",
        "plot_data":             plot_data,
        "iterations":            summary_rows,
        "buffer":                buffer,
        "global_outliers":       global_outliers,
        "outlier_iteration_map": outlier_iteration_map,
    }


def process_rows(df, row_indices, user_tmp_dir, mode):
    """Process rows for outlier detection — iterative removal (mirrors process_column)"""
    results = {}

    for idx, row_num in enumerate(row_indices):
        if row_num >= len(df):
            continue

        row_data       = df.iloc[row_num]
        numeric_values = pd.to_numeric(row_data, errors="coerce").dropna()

        if len(numeric_values) < 3:
            continue

        # ── Iterative outlier removal (same logic as process_column) ──────────
        iteration       = 0
        current         = numeric_values.copy()
        global_outliers = pd.Series(False, index=numeric_values.index)
        # Maps each outlier's label (column name) → 1-based iteration number
        outlier_iteration_map: dict = {}
        summary_rows = []

        while True:
            if len(current) < 3:
                break

            mean_val     = current.mean()
            std_val      = current.std()
            outliers     = (current < mean_val - std_val) | (current > mean_val + std_val)
            num_outliers = outliers.sum()

            summary_rows.append([
                f"Iteration {iteration + 1}",
                round(float(mean_val), 4),
                round(float(std_val),  4),
                int(num_outliers)
            ])

            if num_outliers == 0:
                break

            newly_found = current[outliers].index
            global_outliers.loc[newly_found] = True

            for label in newly_found:
                outlier_iteration_map[label] = iteration + 1   # 1-based

            current   = current[~outliers]
            iteration += 1

        # ── Plot ──────────────────────────────────────────────────────────────
        plt.figure(figsize=(9, 4))

        all_labels = list(numeric_values.index)
        positions  = range(len(all_labels))

        # One colour per bar — red if outlier, blue if normal
        colors = ['red' if global_outliers[lbl] else 'blue' for lbl in all_labels]

        plt.bar(positions, numeric_values.values, color=colors, alpha=0.8)

        # Dummy handles so the legend shows both colours
        from matplotlib.patches import Patch
        legend_handles = [
            Patch(color='blue', alpha=0.8, label='Data'),
            Patch(color='red',  alpha=0.8, label='Outliers'),
        ]

        plt.axhline(y=current.mean(), color="green", linestyle="--",
                    label="Mean (Final)", linewidth=2)

        plt.xticks(positions, all_labels, rotation=45, ha='right', fontsize=8)
        plt.title(f"Row {row_num + 1} - Outlier Detection")
        plt.xlabel("Features")
        plt.ylabel("Values")
        plt.legend(handles=legend_handles + [plt.Line2D([0], [0], color='green',
                    linestyle='--', linewidth=2, label='Mean (Final)')])
        plt.tight_layout()

        buffer = BytesIO()
        plt.savefig(buffer, format="png", dpi=100, bbox_inches='tight')
        buffer.seek(0)
        plot_data = base64.b64encode(buffer.getvalue()).decode()
        plt.close()

        # ── Stats ─────────────────────────────────────────────────────────────
        cleaned_total      = len(current)
        total_values       = len(numeric_values)
        total_outliers     = total_values - cleaned_total
        outlier_percentage = (total_outliers / total_values) * 100 if total_values > 0 else 0

        key = f"Row_{row_num + 1}"
        results[key] = {
            "row_number":            row_num + 1,
            "total_values":          total_values,
            "cleaned_total":         cleaned_total,
            "outlier_count":         total_outliers,
            "outlier_percentage":    round(outlier_percentage, 2),
            "final_mean":            round(float(current.mean()), 4) if cleaned_total > 0 else "N/A",
            "final_std":             round(float(current.std()),  4) if cleaned_total > 0 else "N/A",
            "plot_data":             plot_data,
            "buffer":                buffer,
            "iterations":            summary_rows,
            "global_outliers":       global_outliers,
            "outlier_iteration_map": outlier_iteration_map,   # column label → iteration
            "outliers":              numeric_values[global_outliers].to_dict()
        }

    return results


def create_results_excel(df, results, excel_path, processing_mode, selected_items):
    """Create comprehensive Excel file with all results"""
    df_copy = df.copy()

    df_copy.to_excel(excel_path, sheet_name="Data", index=False)
    book = load_workbook(excel_path)
    ws   = book["Data"]

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                              fill_type="solid")

    # ── Highlight selected columns / rows in yellow ───────────────────────────
    if processing_mode in ["single_column", "multiple_columns"]:
        for col_idx, col_name in enumerate(df_copy.columns, 1):
            if col_name in selected_items:
                ws.cell(row=1, column=col_idx).fill = yellow_fill
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).fill = yellow_fill

    elif processing_mode in ["single_row", "multiple_rows"]:
        for row_num in [int(item) for item in selected_items]:
            excel_row = row_num + 1
            if excel_row <= ws.max_row:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=excel_row, column=col).fill = yellow_fill

    # ── Column-mode: per-iteration green highlighting ─────────────────────────
    if processing_mode in ["single_column", "multiple_columns"]:
        for col_name, result in results.items():
            if result is None:
                continue

            # Find Excel column index for this rate column
            col_idx = None
            for ci, cn in enumerate(df_copy.columns, 1):
                if cn == col_name:
                    col_idx = ci
                    break
            if col_idx is None:
                continue

            for df_idx, iter_num in result.get("outlier_iteration_map", {}).items():
                # df_idx is the DataFrame row index; +2 accounts for header row
                excel_row = df_idx + 2
                color = ITERATION_COLORS[min(iter_num - 1, len(ITERATION_COLORS) - 1)]
                ws.cell(row=excel_row, column=col_idx).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

            # ── Summary sheet ─────────────────────────────────────────────────
            summary = book.create_sheet(f"Summary_{col_name}")
            summary["A1"] = "Iteration"
            summary["B1"] = "Mean"
            summary["C1"] = "Std"
            summary["D1"] = "Outliers Found"

            for row in result["iterations"]:
                summary.append(row)

            summary.append([])
            summary.append(["Total Values",   result["total"]])
            summary.append(["Cleaned Total",  result["cleaned_total"]])
            summary.append(["Total Outliers", result["total_outliers"]])
            summary.append(["Outlier %",      f"{result['outlier_percentage']:.2f}%"])
            summary.append(["Final Mean",     result["final_mean"]])
            summary.append(["Final Std",      result["final_std"]])

            _append_color_legend(summary)

            summary.column_dimensions["A"].width = 20
            summary.column_dimensions["B"].width = 15
            summary.column_dimensions["C"].width = 15
            summary.column_dimensions["D"].width = 18

    # ── Row-mode: per-iteration green highlighting ────────────────────────────
    elif processing_mode in ["single_row", "multiple_rows"]:
        # Build a column-name → Excel-column-index lookup once
        col_name_to_idx = {cn: ci for ci, cn in enumerate(df_copy.columns, 1)}

        for row_key, result in results.items():
            if result is None:
                continue

            row_num   = result["row_number"]          # 1-based
            excel_row = row_num + 1                   # +1 for header

            # Colour each outlier cell with its iteration's shade
            for col_label, iter_num in result.get("outlier_iteration_map", {}).items():
                col_idx = col_name_to_idx.get(col_label)
                if col_idx is None:
                    continue
                color = ITERATION_COLORS[min(iter_num - 1, len(ITERATION_COLORS) - 1)]
                ws.cell(row=excel_row, column=col_idx).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

            # ── Summary sheet ─────────────────────────────────────────────────
            summary = book.create_sheet(f"Summary_{row_key}")
            summary["A1"] = "Iteration"
            summary["B1"] = "Mean"
            summary["C1"] = "Std"
            summary["D1"] = "Outliers Found"

            for row in result["iterations"]:
                summary.append(row)

            summary.append([])
            summary.append(["Row Number",    result["row_number"]])
            summary.append(["Total Values",  result["total_values"]])
            summary.append(["Cleaned Total", result["cleaned_total"]])
            summary.append(["Total Outliers",result["outlier_count"]])
            summary.append(["Outlier %",     f"{result['outlier_percentage']:.2f}%"])
            summary.append(["Final Mean",    result["final_mean"]])
            summary.append(["Final Std",     result["final_std"]])

            if result["outliers"]:
                summary.append([])
                summary.append(["Outlier Values"])
                for col, val in result["outliers"].items():
                    summary.append([col, val])

            _append_color_legend(summary)

            summary.column_dimensions["A"].width = 25
            summary.column_dimensions["B"].width = 20
            summary.column_dimensions["C"].width = 15
            summary.column_dimensions["D"].width = 18

    book.save(excel_path)


def _append_color_legend(ws):
    """Append an iteration color legend at the bottom of a summary worksheet."""
    ws.append([])
    ws.append(["Iteration Color Legend"])
    for i, color in enumerate(ITERATION_COLORS, 1):
        cell = ws.cell(row=ws.max_row + 1, column=1, value=f"Iteration {i}")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


@login_required
def download_result(request, excel_filename):
    tmp_folder     = request.GET.get("tmp_folder")
    selected_items = request.GET.get("selected_items", "")

    if not tmp_folder:
        raise Http404("Temporary folder not specified.")

    safe_path = os.path.normpath(os.path.join(tmp_folder, excel_filename))
    if not os.path.exists(safe_path):
        raise Http404("File not found")

    response = FileResponse(open(safe_path, "rb"), as_attachment=True)
    response["Content-Disposition"] = f'attachment; filename="{excel_filename}"'
    return response