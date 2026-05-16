import os
import io
import json
import uuid  # ADD THIS IMPORT
from urllib.parse import unquote
from django.conf import settings
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.contrib import messages

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from urllib.parse import urljoin

from .forms import TripLengthForm
from .services import (
    build_time_edges, normalize_frequencies,
    save_matrix_to_media, run_octave_script, load_fitting_results,
    get_plot_url, cleanup_session_data
)
import pandas as pd
import numpy as np

from django.contrib.auth.decorators import login_required




def _get_or_create_session_id(request):
    """Get or create a unique session ID"""
    if 'tld_session_id' not in request.session:
        request.session['tld_session_id'] = str(uuid.uuid4())
    return request.session['tld_session_id']



@login_required 

def index(request):
    # Clean up any previous session data
    if 'tld_session_id' in request.session:
        cleanup_session_data(request.session['tld_session_id'])
        del request.session['tld_session_id']
    
    form = TripLengthForm()
    return render(request, "trip_length_distribution/index.html", {"form": form})

def time_entry(request):
    if request.method != "POST":
        return redirect(reverse("trip_length_distribution:manual"))

    # Create session ID
    session_id = _get_or_create_session_id(request)

    # allow commas or dots
    def norm(v):
        return v.replace(",", ".").strip() if isinstance(v, str) else v

    form = TripLengthForm(request.POST)
    errors = {}
    if form.is_valid():
        try:
            minimum = float(norm(form.cleaned_data["Minimum"]))
        except Exception:
            errors["Minimum"] = "Please enter a valid number."

        try:
            maximum = float(norm(form.cleaned_data["Maximum"]))
        except Exception:
            errors["Maximum"] = "Please enter a valid number."

        intervals = form.cleaned_data["intervals"]

        if not errors:
            if intervals < 1:
                errors["intervals"] = "Intervals must be at least 1."
            if maximum <= minimum:
                errors["Maximum"] = "Maximum must be greater than Minimum."

        if errors:
            return render(
                request,
                "trip_length_distribution/index.html",
                {"form": form, "errors": errors},
            )

        # compute edges
        step, array_of_time = build_time_edges(minimum, maximum, intervals)

        # save to session
        request.session["numberOfIntervals"] = intervals
        request.session["minimumTime"] = minimum
        request.session["arrayOfTime"] = array_of_time
        request.session["session_id"] = session_id  # Store session ID

        ctx = {
            "numberOfIntervals": intervals,
            "numberOfIntervals_range": range(intervals),
            "minimumTime": minimum,
            "arrayOfTime": array_of_time,
        }
        return render(request, "trip_length_distribution/time.html", ctx)

    return render(
        request,
        "trip_length_distribution/index.html",
        {"form": form, "errors": {"form": "Please correct the errors."}},
    )


@login_required
def choose_method(request):
    """Render a choice page allowing Manual Entry or Upload Excel."""
    # Clean up any previous session-specific TLD files
    if 'tld_session_id' in request.session:
        try:
            cleanup_session_data(request.session['tld_session_id'])
        except Exception:
            pass
        try:
            del request.session['tld_session_id']
        except Exception:
            pass

    return render(request, "trip_length_distribution/choose_method.html")
@login_required 
def send_to_octave(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Unsupported method"}, status=405)

    # Get session ID
    session_id = request.session.get("session_id")
    if not session_id:
        return JsonResponse({"status": "error", "message": "Session expired"}, status=400)

    # session check
    required = ["arrayOfTime", "numberOfIntervals", "minimumTime"]
    if not all(k in request.session for k in required):
        return JsonResponse({"status": "error", "message": "Session data missing"}, status=400)

    array_of_time = request.session["arrayOfTime"]
    intervals = int(request.session["numberOfIntervals"])

    # collect frequencies
    freqs = []
    for i in range(intervals):
        key = f"time{i}"
        raw = request.POST.get(key, "0")
        try:
            val = float(str(raw).replace(",", "."))
            if val < 0:
                return JsonResponse({"status": "error", "message": f"Frequency {i+1} must be >= 0"}, status=400)
            freqs.append(val)
        except Exception:
            return JsonResponse({"status": "error", "message": f"Invalid frequency at row {i+1}"}, status=400)

    if sum(freqs) <= 0:
        return JsonResponse({"status": "error", "message": "Sum of frequencies must be > 0"}, status=400)

    # save to session
    request.session["frequencies"] = freqs

    # normalize + write json
    normed = normalize_frequencies(freqs)
    save_matrix_to_media(array_of_time, normed, session_id)

    # run octave
    try:
        run_octave_script(session_id)
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

    return redirect(reverse("trip_length_distribution:output"))


@login_required
def upload_excel_start(request):
    """Step 1: Upload Excel file and show available columns."""
    if request.method != "POST":
        return render(request, "trip_length_distribution/upload_excel_start.html")

    # Get file
    f = request.FILES.get("excel_file")
    if not f:
        messages.error(request, "No file uploaded.")
        return render(request, "trip_length_distribution/upload_excel_start.html")

    # Create session ID if needed
    session_id = _get_or_create_session_id(request)
    request.session["session_id"] = session_id

    try:
        df = pd.read_excel(f)
    except Exception as e:
        messages.error(request, f"Failed to read Excel file: {e}")
        return render(request, "trip_length_distribution/upload_excel_start.html")

    if df.empty:
        messages.error(request, "Excel file is empty.")
        return render(request, "trip_length_distribution/upload_excel_start.html")

    # Get all columns
    columns = df.columns.tolist()
    
    # Try to find numeric columns
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    
    # Store dataframe in session (as dict/JSON for serialization)
    request.session["excel_data"] = df.to_json()
    request.session["excel_columns"] = columns
    request.session["numeric_columns"] = numeric_cols

    return render(request, "trip_length_distribution/upload_excel_select_column.html", {
        "columns": columns,
        "numeric_columns": numeric_cols,
    })


@login_required
def upload_excel_select_column(request):
    """Step 2: User selects which column is the time data."""
    if request.method != "POST":
        # Show selection form
        columns = request.session.get("excel_columns", [])
        numeric_cols = request.session.get("numeric_columns", [])
        return render(request, "trip_length_distribution/upload_excel_select_column.html", {
            "columns": columns,
            "numeric_columns": numeric_cols,
        })

    selected_col = request.POST.get("time_column")
    if not selected_col:
        messages.error(request, "Please select a column.")
        columns = request.session.get("excel_columns", [])
        numeric_cols = request.session.get("numeric_columns", [])
        return render(request, "trip_length_distribution/upload_excel_select_column.html", {
            "columns": columns,
            "numeric_columns": numeric_cols,
        })

    # Load dataframe from session
    excel_json = request.session.get("excel_data")
    if not excel_json:
        messages.error(request, "Session data lost. Please upload again.")
        return redirect(reverse("trip_length_distribution:upload_excel"))

    try:
        df = pd.read_json(excel_json)
    except Exception as e:
        messages.error(request, f"Failed to read session data: {e}")
        return redirect(reverse("trip_length_distribution:upload_excel"))

    # Extract numeric data from selected column
    try:
        series = pd.to_numeric(df[selected_col], errors="coerce").dropna()
        if series.empty:
            messages.error(request, f"No numeric data found in column '{selected_col}'.")
            columns = request.session.get("excel_columns", [])
            numeric_cols = request.session.get("numeric_columns", [])
            return render(request, "trip_length_distribution/upload_excel_select_column.html", {
                "columns": columns,
                "numeric_columns": numeric_cols,
            })
    except Exception as e:
        messages.error(request, f"Error processing column: {e}")
        return redirect(reverse("trip_length_distribution:upload_excel"))

    # Calculate min, max, suggested intervals
    min_val = float(series.min())
    max_val = float(series.max())
    data_count = len(series)
    suggested_intervals = max(5, int(np.sqrt(data_count)))

    # Store in session
    request.session["time_data"] = series.tolist()
    request.session["selected_column"] = selected_col
    request.session["data_min"] = min_val
    request.session["data_max"] = max_val

    return render(request, "trip_length_distribution/upload_excel_generate_intervals.html", {
        "data_min": min_val,
        "data_max": max_val,
        "data_count": data_count,
        "suggested_intervals": suggested_intervals,
    })


@login_required
def upload_excel_generate_intervals(request):
    """Step 3: User specifies number of intervals and sees the breakdown."""
    if request.method == "GET":
        min_val = request.session.get("data_min", 0)
        max_val = request.session.get("data_max", 100)
        data_count = len(request.session.get("time_data", []))
        suggested = max(5, int(np.sqrt(data_count)))
        return render(request, "trip_length_distribution/upload_excel_generate_intervals.html", {
            "data_min": min_val,
            "data_max": max_val,
            "data_count": data_count,
            "suggested_intervals": suggested,
        })

    # POST: user submitted interval count
    intervals_str = request.POST.get("intervals", "0")
    try:
        intervals = int(intervals_str)
    except ValueError:
        messages.error(request, "Please enter a valid number of intervals.")
        return redirect(reverse("trip_length_distribution:upload_excel_generate_intervals"))

    if intervals < 1:
        messages.error(request, "Intervals must be at least 1.")
        return redirect(reverse("trip_length_distribution:upload_excel_generate_intervals"))

    # Get time data from session
    time_data = request.session.get("time_data", [])
    min_val = request.session.get("data_min")
    max_val = request.session.get("data_max")

    if not time_data:
        messages.error(request, "Session data lost. Please upload again.")
        return redirect(reverse("trip_length_distribution:upload_excel"))

    # Generate intervals using the same logic as build_time_edges
    step, array_of_time = build_time_edges(min_val, max_val, intervals)

    # Build bins and compute frequencies
    bins = [min_val] + array_of_time
    try:
        series = pd.Series(time_data)
        cats = pd.cut(series, bins=bins, right=False)
        counts = list(cats.value_counts(sort=False).astype(float).tolist())
    except Exception as e:
        messages.error(request, f"Failed to compute frequencies: {e}")
        return redirect(reverse("trip_length_distribution:upload_excel_generate_intervals"))

    # Prepare display data
    interval_data = []
    for i, (edge, count) in enumerate(zip(array_of_time, counts)):
        left = array_of_time[i-1] if i > 0 else min_val
        right = edge
        interval_data.append({
            "interval": f"{left:.2f} - {right:.2f}",
            "count": int(count),
        })

    # Store in session
    request.session["intervals"] = intervals
    request.session["numberOfIntervals"] = intervals
    request.session["arrayOfTime"] = array_of_time
    request.session["minimumTime"] = min_val
    request.session["frequencies"] = [float(x) for x in counts]

    return render(request, "trip_length_distribution/upload_excel_review_intervals.html", {
        "interval_data": interval_data,
        "total_values": sum(counts),
    })


@login_required
def upload_excel_process(request):
    """Step 4: Process the intervals and run Octave."""
    session_id = request.session.get("session_id")
    frequencies = request.session.get("frequencies", [])
    array_of_time = request.session.get("arrayOfTime", [])

    if not session_id or not frequencies or not array_of_time:
        messages.error(request, "Session incomplete. Please upload again.")
        return redirect(reverse("trip_length_distribution:upload_excel"))

    try:
        # Normalize, save matrix, run octave
        normed = normalize_frequencies(frequencies)
        save_matrix_to_media(array_of_time, normed, session_id)
        run_octave_script(session_id)
    except Exception as e:
        messages.error(request, f"Processing failed: {e}")
        return redirect(reverse("trip_length_distribution:upload_excel"))

    return redirect(reverse("trip_length_distribution:output"))


@login_required
def upload_excel_time(request):
    """Redirect to new upload flow."""
    return redirect(reverse("trip_length_distribution:upload_excel"))




@login_required  
def output(request):
    # Get session ID
    session_id = request.session.get("session_id")
    if not session_id:
        messages.warning(request, "Session expired. Please start over.")
        return redirect("trip_length_distribution:manual")

    # Guard: make sure we have session data first
    number_of_intervals = request.session.get("numberOfIntervals")
    frequencies = request.session.get("frequencies")

    if not number_of_intervals or not frequencies:
        messages.warning(request, "Please start by entering Trip Length Distribution inputs.")
        return redirect("trip_length_distribution:manual")

    data = {
        "numberOfIntervals": number_of_intervals,
        "frequencies": frequencies,
        "totalFrequencySum": sum(frequencies),
        "session_id": session_id,  # Include session ID in data
    }

    # Load fitting_results.json if present
    fitting_results = load_fitting_results(session_id)
    if fitting_results:
        data["fittingResults"] = fitting_results

    # URL for the generated plot (session-specific)
    plot_url = get_plot_url(session_id)

    return render(
        request,
        "trip_length_distribution/output.html",
        {"data": data, "plot_url": plot_url},
    )
@login_required
def download_pdf(request):
    data_param = request.GET.get("data")
    if not data_param:
        return HttpResponse("No data provided.", status=400)

    try:
        decoded = unquote(data_param)
        data = json.loads(decoded)
    except Exception as e:
        return HttpResponse(f"Invalid data: {e}", status=400)

    # Get session ID from data or request
    session_id = data.get("session_id") or request.session.get("session_id")
    if not session_id:
        return HttpResponse("Session expired.", status=400)

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width/2, height - 20, "Fitting Results Report")

    pdf.setFont("Helvetica", 12)
    y = height - 80
    pdf.drawString(50, y, f"Number of Intervals: {data.get('numberOfIntervals', 0)}")

    total_freq = sum(data.get('frequencies', [])) or 0
    table_data = [["Interval", "Frequency", "Normalized"]]
    inputs = data.get("fittingResults", {}).get("inputs", {})
    t = inputs.get("time", [])
    f = inputs.get("frequency", [])

    t = inputs.get("time", [])

    # نحسب الفرق بين أول قيمتين (لو فيه أكتر من قيمة)
    if len(t) > 1:
        diff = t[1] - t[0]
    else:
        diff = 0  # لو فيه قيمة واحدة بس هنخليها تبدأ من نفس الرقم

    for i, freq in enumerate(data.get("frequencies", [])):
        if i == 0:
            left = t[i] - diff   # ⬅️ هنا التعديل المهم
            right = t[i]
            label = f"{left} - {right}"
        else:
            left = t[i-1]
            right = t[i]
            label = f"{left} - {right}"

        norm = (freq / total_freq) if total_freq else 0
        table_data.append([label, f"{freq:.2f}", f"{norm:.4f}"])



    table_data.append(["Total", f"{sum(data.get('frequencies', [])):.4f}", "100%"])

    table = Table(table_data, colWidths=[200, 150, 150])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    table.wrapOn(pdf, 50, height-500)
    table.drawOn(pdf, 50, height-500)

    # Use session-specific image path
    image_path = os.path.join(settings.MEDIA_ROOT, "TLD", f"session_{session_id}", "Fitting_Results.png")
    if os.path.exists(image_path):
        try:
            pdf.drawImage(image_path, 50, height-760, width=550, height=250, preserveAspectRatio=True, mask='auto')
        except Exception as e:
            pdf.setFont("Helvetica", 10)
            pdf.drawString(50, height-530, f"Error loading image: {str(e)}")
    else:
        pdf.setFont("Helvetica", 10)
        pdf.drawString(50, height-530, "Image not found.")

    if "fittingResults" in data:
        pdf.showPage()
        params = data["fittingResults"]["optimized_parameters"]
        stats = data["fittingResults"]["statistics"]

        y0 = height - 100
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y0, "Optimized Parameters:")
        pdf.setFont("Helvetica", 12)
        pdf.drawString(60, y0-20, f"A = {params['a']:.4f}")
        pdf.drawString(60, y0-40, f"B = {params['b']:.4f}")
        pdf.drawString(60, y0-60, f"C = {params['c']:.4f}")

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y0-90, "Equation:")
        pdf.setFont("Courier", 12)
        pdf.drawString(60, y0-110, f"f(t) = {params['a']:.3f} * t^{params['b']:.3f} * e^({params['c']:.3f} * t)")

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y0-140, "Statistics:")
        pdf.setFont("Helvetica", 12)
        pdf.drawString(60, y0-160, f"R² = {stats['R2']:.4f}")
        pdf.drawString(60, y0-180, f"SSE = {stats['SSE']:.4f}")

    pdf.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type="application/pdf")
@login_required
def download_excel(request):
    data_param = request.GET.get("data")
    if not data_param:
        return HttpResponse("No data provided.", status=400)

    try:
        decoded = unquote(data_param)
        data = json.loads(decoded)
    except Exception as e:
        return HttpResponse(f"Invalid data: {e}", status=400)

    # Get session ID from data or request
    session_id = data.get("session_id") or request.session.get("session_id")
    if not session_id:
        return HttpResponse("Session expired.", status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fitting Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="8b0000")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_align = Alignment(horizontal="center", vertical="center")

    ws.append(["Interval", "Frequency", "Normalized"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    total_freq = sum(data.get("frequencies", [])) or 0
    inputs = data.get("fittingResults", {}).get("inputs", {})
    t = inputs.get("time", [])

    # نحسب الفرق بين أول قيمتين (لو فيه أكتر من قيمة)
    if len(t) > 1:
        diff = t[1] - t[0]
    else:
        diff = 0  # لو فيه قيمة واحدة بس هنخليها تبدأ من نفس الرقم

    for i, freq in enumerate(data.get("frequencies", [])):
        if i == 0:
            left = t[i] - diff   # ⬅️ هنا التعديل المهم
            right = t[i]
            label = f"{left} - {right}"
        else:
            left = t[i-1]
            right = t[i]
            label = f"{left} - {right}"

        norm = (freq / total_freq) if total_freq else 0
        ws.append([label, freq, norm])


    ws.append(["Total", sum(data.get("frequencies", [])), 1 if total_freq else 0])
    # style last row
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="000000")

    ws.append([])
    ws.append([])
    ws.append(["Parameter", "Value"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill

    if "fittingResults" in data:
        params = data["fittingResults"]["optimized_parameters"]
        stats = data["fittingResults"]["statistics"]
        ws.append(["Optimized Parameters"])
        for cell in ws[ws.max_row]:
            cell.font = bold_font

        ws.append(["A", params["a"]])
        ws.append(["B", params["b"]])
        ws.append(["C", params["c"]])

        ws.append([])
        ws.append(["Statistics"])
        for cell in ws[ws.max_row]:
            cell.font = bold_font

        ws.append(["R²", stats["R2"]])
        ws.append(["SSE", stats["SSE"]])

        ws.append([])
        ws.append(["Equation", f"f(t) = {params['a']:.3f} * t^{params['b']:.3f} * e^({params['c']:.3f} * t)"])

    # Use session-specific image path
    img_path = os.path.join(settings.MEDIA_ROOT, "TLD", f"session_{session_id}", "Fitting_Results.png")
    if os.path.exists(img_path):
        try:
            img = ExcelImage(img_path)
            img.width = 600
            img.height = 400
            ws.add_image(img, "E2")
        except Exception:
            ws.append(["Image could not be loaded."])
    else:
        ws.append(["Image not found."])

    # borders & alignment
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align

    # autosize
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return HttpResponse(
        buf,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    
    
    